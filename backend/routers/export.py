"""
Export Utilities
Funzioni per esportare dati in vari formati (Excel, CSV, PDF)
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from typing import List, Optional
import io
import logging
from datetime import datetime

from auth import User, get_current_active_user
from core.database import get_db

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/export", tags=["export"])


@router.get("/properties/excel")
async def export_properties_excel(
    status: Optional[str] = None,
    location: Optional[str] = None,
    current_user: User = Depends(get_current_active_user),
    db = Depends(get_db)
):
    """
    Esporta properties in formato Excel
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Librerie export non installate. Esegui: pip install pandas openpyxl"
        )
    
    logger.info(f"Exporting properties to Excel (user: {current_user.username})")
    
    # Build query
    query = {}
    if current_user.role == "agent":
        query['agent_id'] = current_user.id
    
    if status:
        query['status'] = status
    
    if location:
        query['location'] = {'$regex': location, '$options': 'i'}
    
    # Fetch properties
    properties = await db.properties.find(query, {"_id": 0}).to_list(1000)
    
    if not properties:
        raise HTTPException(status_code=404, detail="Nessun immobile trovato")
    
    # Convert to DataFrame
    df = pd.DataFrame(properties)
    
    # Select and rename columns
    columns_map = {
        'reference': 'Codice',
        'title': 'Titolo',
        'property_type': 'Tipo',
        'location': 'Zona',
        'price': 'Prezzo (€)',
        'square_meters': 'Superficie (mq)',
        'bedrooms': 'Camere',
        'bathrooms': 'Bagni',
        'status': 'Stato',
        'created_at': 'Data Creazione'
    }
    
    # Keep only desired columns
    available_cols = [col for col in columns_map.keys() if col in df.columns]
    df = df[available_cols]
    df = df.rename(columns=columns_map)
    
    # Format datetime
    if 'Data Creazione' in df.columns:
        df['Data Creazione'] = pd.to_datetime(df['Data Creazione']).dt.strftime('%d/%m/%Y')
    
    # Format price
    if 'Prezzo (€)' in df.columns:
        df['Prezzo (€)'] = df['Prezzo (€)'].apply(lambda x: f"€ {x:,.0f}")
    
    # Create Excel with styling
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Immobili', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Immobili']
        
        # Style header
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"immobili_tempocasa_{timestamp}.xlsx"
    
    logger.info(f"Excel export completed: {len(properties)} properties")
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/clients/excel")
async def export_clients_excel(
    current_user: User = Depends(get_current_active_user),
    db = Depends(get_db)
):
    """
    Esporta clienti in formato Excel
    """
    try:
        import pandas as pd
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Pandas non installato. Esegui: pip install pandas openpyxl"
        )
    
    logger.info(f"Exporting clients to Excel (user: {current_user.username})")
    
    # Fetch clients
    clients = await db.clients.find({}, {"_id": 0}).to_list(1000)
    
    if not clients:
        raise HTTPException(status_code=404, detail="Nessun cliente trovato")
    
    # Convert to DataFrame
    df = pd.DataFrame(clients)
    
    # Select columns
    columns_map = {
        'name': 'Nome',
        'surname': 'Cognome',
        'phone': 'Telefono',
        'email': 'Email',
        'looking_for': 'Cerca',
        'budget': 'Budget (€)',
        'profile_completed': 'Profilo Completo',
        'created_at': 'Data Registrazione'
    }
    
    available_cols = [col for col in columns_map.keys() if col in df.columns]
    df = df[available_cols]
    df = df.rename(columns=columns_map)
    
    # Format datetime
    if 'Data Registrazione' in df.columns:
        df['Data Registrazione'] = pd.to_datetime(df['Data Registrazione']).dt.strftime('%d/%m/%Y')
    
    # Format budget
    if 'Budget (€)' in df.columns:
        df['Budget (€)'] = df['Budget (€)'].apply(
            lambda x: f"€ {x:,.0f}" if pd.notna(x) else "Non specificato"
        )
    
    # Create Excel
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Clienti', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Clienti']
        
        # Style header
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"clienti_tempocasa_{timestamp}.xlsx"
    
    logger.info(f"Excel export completed: {len(clients)} clients")
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/properties/csv")
async def export_properties_csv(
    status: Optional[str] = None,
    current_user: User = Depends(get_current_active_user),
    db = Depends(get_db)
):
    """
    Esporta properties in formato CSV
    """
    try:
        import pandas as pd
    except ImportError:
        raise HTTPException(status_code=500, detail="Pandas non installato")
    
    # Build query
    query = {}
    if current_user.role == "agent":
        query['agent_id'] = current_user.id
    if status:
        query['status'] = status
    
    # Fetch properties
    properties = await db.properties.find(query, {"_id": 0}).to_list(1000)
    
    if not properties:
        raise HTTPException(status_code=404, detail="Nessun immobile trovato")
    
    # Convert to DataFrame and CSV
    df = pd.DataFrame(properties)
    
    output = io.StringIO()
    df.to_csv(output, index=False, encoding='utf-8-sig')
    output.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"immobili_tempocasa_{timestamp}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
